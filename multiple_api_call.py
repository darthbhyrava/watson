import pandas as pd
import openpyxl
import time
from watson_developer_cloud import NaturalLanguageUnderstandingV1
from watson_developer_cloud.natural_language_understanding_v1 \
  import Features, SentimentOptions

natural_language_understanding = NaturalLanguageUnderstandingV1(username='fe7e83ab-fe27-432a-816e-428e79c992db',password='kaTY4H6r3HbY',version='2017-02-27')

wb = openpyxl.load_workbook('21Feb_Pos.xlsx')
ws = wb.active
pred_sent_col = ws['L']
man_sent_col = ws['Q']
text_col = ws['U']
rows = len(pred_sent_col)

correct_p = 0.0
all_rows = []
for row in range(1,rows):
    print row
    time.sleep(0.1)
    towrite = {}
    pred_sent = pred_sent_col[row].value
    man_sent = man_sent_col[row].value
    text = text_col[row].value.replace(';','')
    try:
        response = natural_language_understanding.analyze(text=text, features=Features(sentiment=SentimentOptions(document=True)))
    except: 
        continue
    label = response['sentiment']['document']['label']
    score = response['sentiment']['document']['score']
    towrite['Predicted Sentiment'] = pred_sent
    towrite['Manual Sentiment'] = man_sent
    towrite['Watson_Label'] = label
    if label == 'positive':
        correct_p+=1.0
    towrite['Watson_Score'] = score
    print towrite
    all_rows.append(towrite)
percentage = correct_p/float(len(pred_sent_col))*100
print percentage
df = pd.DataFrame(all_rows)
df.to_csv("21Feb_Pos.csv")
